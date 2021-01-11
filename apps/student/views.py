import os
import json
import uuid
import hashlib

from django.conf import settings
from django.db.models import Q
from django.shortcuts import render
from django.http import JsonResponse
from student.models import Student

import openpyxl


def get_students(request):
    """获取所有学生信息"""
    # 使用 ORM 获取所有学生信息
    try:
        # 使用 ORM 获取所有学生信息并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "获取学生信息出现异常，错误信息：" + str(e)})


def query_students(request):
    """查询学生信息"""
    # 接收传递过来的查询条件：axios 默认是 json 格式 字典类型('inputstr') data['inputstr']
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_students = Student.objects.filter(Q(sno__icontains=data['inputstr']) |
                                              Q(name__icontains=data['inputstr']) |
                                              Q(gender__icontains=data['inputstr']) |
                                              Q(mobile__icontains=data['inputstr']) |
                                              Q(email__icontains=data['inputstr']) |
                                              Q(address__icontains=data['inputstr'])).values()
        # 把结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "查询学生信息出现异常，错误信息：" + str(e)})


def is_exists_sno(request):
    """判断学号是否存在"""
    # 接受传来的学号
    data = json.loads(request.body.decode('utf-8'))
    # 进行校验
    try:
        obj_students = Student.objects.filter(sno=data['sno'])
        if obj_students.count() == 0:
            return JsonResponse({'code': 1, 'exists': False})
        else:
            return JsonResponse({'code': 1, 'exists': True})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '校验学号失败，错误信息为：' + str(e)})


def add_student(request):
    """添加学生到数据库"""
    # 接受全段传过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 添加到数据库
        obj_student = Student(sno=data['sno'], name=data['name'], gender=data['gender'],
                              birthday=data['birthday'], mobile=data['mobile'],
                              email=data['email'], address=data['address'], image=data['image'])
        # 执行添加操作
        obj_student.save()
        # 使用 ORM 获取所有学生信息并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "添加到数据库出现异常，具体原因：" + str(e)})


def update_student(request):
    """修改学生到数据库"""
    # 接受全段传过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 查找要更新的学生信息
        obj_student = Student.objects.get(sno=data['sno'])
        # 依次修改
        obj_student.name = data['name']
        obj_student.gender = data['gender']
        obj_student.birthday = data['birthday']
        obj_student.mobile = data['mobile']
        obj_student.email = data['email']
        obj_student.address = data['address']
        obj_student.image = data['image']
        # 保存
        obj_student.save()
        # 使用 ORM 获取所有学生信息并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "修改保存到数据库出现异常，具体原因：" + str(e)})


def delete_student(request):
    """删除学生到数据库"""
    # 接收前段传过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 查找要更新的学生信息
        obj_student = Student.objects.get(sno=data['sno'])
        # 删除
        obj_student.delete()
        # 使用 ORM 获取所有学生信息并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "删除学生信息时出现异常，具体原因：" + str(e)})


def delete_students(request):
    """批量删除学生到数据库"""
    # 接收前段传过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 遍历传递来的集合
        for one_student in data['student']:
            # 查询当前记录
            obj_student = Student.objects.get(sno=one_student['sno'])
            # 执行删除
            obj_student.delete()
        # 使用 ORM 获取所有学生信息并把对象转为字典格式
        obj_students = Student.objects.all().values()
        # 把外层结果转为 list
        students = list(obj_students)
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "批量删除学生信息时出现异常，具体原因：" + str(e)})


def upload(request):
    """接收上传的文件"""
    # 接收上传的文件(缓存中)
    rev_file = request.FILES.get('avatar')  # 前段的名字 avatar
    # 判断是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '图片不存在！'})
    # 获得唯一的名字：uuid + hash
    new_name = get_random_str()
    # 准备写入的 url
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])  # + 图片后缀
    # 开始写入到磁盘
    try:
        f = open(file_path, 'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        f.close()
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(rev_file.name)[1]})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


def get_random_str():
    """生成随即字符串"""
    # 获取 uuid 的随机数
    uuid_val = uuid.uuid4()
    # 获取 uuid 的随即数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    # 获取 md5 实例
    md5 = hashlib.md5()
    # 拿取 uuid 的 md5 摘要
    md5.update(uuid_str)
    # 返回固定长度的字符串
    return md5.hexdigest()


def read_excel_dict(path:str):
    """读取Excel数据存储为包含字典的列表"""
    # 实例化一个 wb
    workbook = openpyxl.load_workbook(path)
    # 实例化 sheet
    sheet = workbook['student']
    # 定义一个变量存储最终数据 -- []
    students = []
    # 准备 key
    keys = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address']
    # 遍历
    for row in sheet.rows:
        # 定义临时用的字典
        temp_dict = {}
        # 组合值和 key
        for index, cell in enumerate(row):
            temp_dict[keys[index]] = cell.value
        # 添加到 list 中
        students.append(temp_dict)
    return students


def import_students_excel(request):
    """从Excel批量导入学生信息"""
    # 1. 接收Excel文件存储到后端 media 文件
    rev_file = request.FILES.get('excel')
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '文件不存在！'})
    new_name = get_random_str()
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})
    # 2. 读取存储在 media 文件的数据
    ex_students = read_excel_dict(file_path)
    # 3. 读取数据存储到数据库
    # 定义变量success error
    success = 0
    error = 0
    error_snos = []
    # 开始遍历
    for one_student in ex_students:
        try:
            # obj_student = Student(sno=one_student['sno'],
            # 上面默认是更新+创建所以重复导入不会报错会更新，可以用下面的
            obj_student = Student.objects.create(sno=one_student['sno'],
                                                 name=one_student['name'],
                                                 gender=one_student['gender'],
                                                 birthday=one_student['birthday'],
                                                 mobile=one_student['mobile'],
                                                 email=one_student['email'],
                                                 address=one_student['address'])
            obj_student.save()  # 保存
            success += 1  # 计数
        except:
            error += 1
            error_snos.append(one_student['sno'])
    # 4. 返回要导入的所有学生信息（成功条数，失败的学号）
    obj_student = Student.objects.all().values()
    students = list(obj_student)
    return JsonResponse({'code': 1, 'success': success, 'error': error, 'errors': error_snos, 'data': students})


def export_student_excel(request):
    """导出数据到excel"""
    # 获取所有的学生信息
    obj_students = Student.objects.all().values()
    # 转为 list
    students = list(obj_students)
    # 定义文件名
    excel_name = get_random_str() + '.xlsx'
    # 准备写入的路径
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    write_to_excel(students, path)
    # 返回
    return JsonResponse({'code': 1, 'name': excel_name})


def write_to_excel(data:list, path:str):
    """导出数据库数据并写入到excel"""
    # 实例化 workbook
    workbook = openpyxl.Workbook()
    # 激活一个 sheet
    sheet = workbook.active
    # 命名 sheet
    sheet.title = 'student'
    # 准备 keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)
